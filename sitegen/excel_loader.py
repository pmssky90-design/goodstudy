from __future__ import annotations

from collections.abc import Iterator
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class SourceError(RuntimeError):
    pass


def load_sheets(path: Path) -> tuple[dict[str, list[dict[str, Any]]], list[dict[str, Any]]]:
    if not path.is_file():
        raise SourceError(f"필요한 입력 엑셀이 없습니다: {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets: dict[str, list[dict[str, Any]]] = {}
    report: list[dict[str, Any]] = []
    try:
        for worksheet in workbook.worksheets:
            iterator: Iterator[tuple[Any, ...]] = worksheet.iter_rows(values_only=True)
            raw_headers = next(iterator, ())
            headers = [str(value).strip() if value is not None else f"_empty_{index}" for index, value in enumerate(raw_headers)]
            rows: list[dict[str, Any]] = []
            for source_row, values in enumerate(iterator, 2):
                if not any(value is not None and str(value).strip() for value in values):
                    continue
                rows.append({headers[i]: values[i] if i < len(values) else None for i in range(len(headers))})
                rows[-1]["_source_row"] = source_row
            sheets[worksheet.title] = rows
            report.append({"sheet": worksheet.title, "rows": len(rows), "columns": [h for h in headers if not h.startswith("_empty_")]})
    finally:
        workbook.close()
    return sheets, report
