from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl


def clean(value: object) -> str:
    return " ".join(str(value or "").split())


def inspect(path: Path) -> dict[str, object]:
    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for sheet in book.worksheets:
        sample = list(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 12), values_only=True))
        best = []
        best_index = 0
        for index, row in enumerate(sample, 1):
            values = [clean(x) for x in row]
            score = sum(any(key in value for key in ("학교", "주소", "시도", "시군구", "교육청")) for value in values)
            if score > sum(any(key in value for key in ("학교", "주소", "시도", "시군구", "교육청")) for value in best):
                best, best_index = values, index
        headers = [x for x in best if x][:50]
        joined = " ".join(headers)
        sheets.append({
            "sheet_name": sheet.title,
            "rows": sheet.max_row,
            "columns": sheet.max_column,
            "header_row": best_index,
            "headers": headers,
            "has_school_name": any("학교명" in x or x in ("학교", "학교명칭") for x in headers),
            "has_region_or_address": any(x in joined for x in ("주소", "시도", "시군구", "교육청")),
        })
    book.close()
    info = path.stat()
    return {
        "path": str(path),
        "modified": info.st_mtime,
        "size": info.st_size,
        "sheets": sheets,
    }


print(json.dumps([inspect(Path(arg)) for arg in sys.argv[1:]], ensure_ascii=False, indent=2))
