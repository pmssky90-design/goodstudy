from __future__ import annotations

import csv
import html
import json
import os
import re
import shutil
import unicodedata
from collections import Counter, defaultdict, deque
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "candidate_output_school_excel_linkfix"
PRIMARY_TARGET = ROOT / "candidate_output_home_region_school_only"
REBUILD_TARGET = ROOT / "candidate_output_home_region_school_only_rebuild"
EXCEL = Path(r"C:\gptwp\자료\2025년 유초중등 학교별 학년별 학생수 학급수 입학 졸업 교원 직원 면적_260206W.xlsx")
SHEET = "학교별 주요통계"
META = ROOT / "intermediate" / "normalized-pages.json"
CHECKPOINT = ROOT / "intermediate" / "home-region-school-only-build.json"
MATCH_CSV = ROOT / "audit" / "school-excel-final-match.csv"
UNMATCHED_CSV = ROOT / "audit" / "school-excel-final-unmatched.csv"
CARD_CSV = ROOT / "audit" / "home-school-card-list.csv"
REMOVAL_CSV = ROOT / "audit" / "home-removal-log.csv"

PROVINCES = [
    ("서울특별시", "seoul"), ("부산광역시", "busan"), ("대구광역시", "daegu"),
    ("인천광역시", "incheon"), ("광주광역시", "gwangju"), ("대전광역시", "daejeon"),
    ("울산광역시", "ulsan"), ("세종특별자치시", "sejong"), ("경기도", "gyeonggi"),
    ("강원특별자치도", "gangwon"), ("충청북도", "chungbuk"), ("충청남도", "chungnam"),
    ("전북특별자치도", "jeonbuk"), ("전라남도", "jeonnam"), ("경상북도", "gyeongbuk"),
    ("경상남도", "gyeongnam"), ("제주특별자치도", "jeju"),
]
PROVINCE_MAP = {
    "서울": "서울특별시", "부산": "부산광역시", "대구": "대구광역시", "인천": "인천광역시",
    "광주": "광주광역시", "대전": "대전광역시", "울산": "울산광역시", "세종": "세종특별자치시",
    "경기": "경기도", "강원": "강원특별자치도", "충북": "충청북도", "충남": "충청남도",
    "전북": "전북특별자치도", "전남": "전라남도", "경북": "경상북도", "경남": "경상남도",
    "제주": "제주특별자치도",
}


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def norm(value: object) -> str:
    value = unicodedata.normalize("NFKC", clean(value)).lower()
    return re.sub(r"[^0-9a-z가-힣]", "", value)


def school_level(value: str) -> str:
    if "초등학교" in value:
        return "초등학교"
    if "중학교" in value:
        return "중학교"
    if "고등학교" in value:
        return "고등학교"
    return "기타"


def address_core(value: object) -> str:
    value = clean(value).split("(")[0]
    value = re.sub(r"\b(?:초등학교|중학교|고등학교)\b.*$", "", value)
    return norm(value)


def district_compatible(site_district: str, excel_district: str) -> bool:
    left, right = norm(site_district), norm(excel_district)
    return bool(left and right and (left == right or left in right or right in left))


def site_name_form(excel_name: str) -> str:
    value = norm(excel_name)
    if value.endswith("고등학교"):
        return value.removesuffix("고등학교") + "고"
    return value


def choose_target() -> Path:
    if PRIMARY_TARGET.exists() and any(PRIMARY_TARGET.iterdir()):
        if REBUILD_TARGET.exists() and any(REBUILD_TARGET.iterdir()):
            raise RuntimeError(f"both target paths are non-empty: {PRIMARY_TARGET}, {REBUILD_TARGET}")
        return REBUILD_TARGET
    return PRIMARY_TARGET


def load_excel() -> list[dict[str, str]]:
    book = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)
    sheet = book[SHEET]
    headers = [clean(x) for x in next(sheet.iter_rows(min_row=17, max_row=17, values_only=True))]
    rows = []
    for row_number, values in enumerate(sheet.iter_rows(min_row=18, values_only=True), 18):
        item = {headers[index]: clean(value) for index, value in enumerate(values) if index < len(headers)}
        if not item.get("학교명"):
            continue
        item.update({
            "_row": str(row_number),
            "_province": PROVINCE_MAP.get(item.get("시도", ""), item.get("시도", "")),
            "_district": item.get("행정구", ""),
            "_level": school_level(item.get("학교급", "")),
        })
        rows.append(item)
    book.close()
    return rows


def match_site(excel_rows: list[dict[str, str]], metadata: list[dict[str, object]]) -> tuple[list[dict[str, object]], list[dict[str, str]]]:
    by_province: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in excel_rows:
        by_province[row["_province"]].append(row)
    general_pages = [x for x in metadata if x.get("node_type") == "school_general"]
    matches, unmatched = [], []
    for site in general_pages:
        province, district = str(site.get("province", "")), str(site.get("city", "") or site.get("district", ""))
        candidates = [
            row for row in by_province.get(province, [])
            if district_compatible(district, row["_district"])
        ]
        site_name, site_address = norm(site.get("school_name", "")), norm(site.get("school_address", ""))
        site_address_core = address_core(site.get("school_address", ""))
        address_matches = [
            row for row in candidates
            if len(site_address_core) >= 12
            and (
                site_address_core in address_core(row["주소"])
                or address_core(row["주소"]) in site_address_core
                or site_address_core[:20] == address_core(row["주소"])[:20]
            )
        ]
        name_region_matches = [
            row for row in candidates
            if site_name_form(row["학교명"]) == site_name
            and row["_level"] == "고등학교"
        ]
        exact = [
            row for row in candidates
            if norm(row["학교명"]) == site_name
            or (
                row in address_matches
                and (
                    norm(row["학교명"]).removesuffix("고등학교") + "고"
                    if norm(row["학교명"]).endswith("고등학교") else norm(row["학교명"])
                ) == site_name
            )
        ]
        pool = exact or name_region_matches or address_matches
        unique = {row["_row"]: row for row in pool}
        if len(unique) != 1:
            unmatched.append({
                "site_school_name": str(site.get("school_name", "")), "site_province": province,
                "site_sigungu": district, "site_address": str(site.get("school_address", "")),
                "site_slug": str(site["slug"]), "candidate_count": str(len(unique)),
                "reason": "ambiguous" if unique else "no_excel_match",
            })
            continue
        row = next(iter(unique.values()))
        reason = (
            "exact_name_region_address" if row in exact
            else ("normalized_name_region_level" if row in name_region_matches else "address_corrected")
        )
        matches.append({"excel": row, "site": site, "match_type": reason})
    return matches, unmatched


def balanced_section(source: str, marker: str) -> tuple[int, int, str]:
    start = source.index(marker)
    token = re.compile(r"<section\b|</section>", re.I)
    depth = 0
    for match in token.finditer(source, start):
        if match.group(0).lower().startswith("<section"):
            depth += 1
        else:
            depth -= 1
            if depth == 0:
                return start, match.end(), source[start:match.end()]
    raise RuntimeError(f"unclosed section: {marker}")


def select_home(matches: list[dict[str, object]]) -> dict[str, list[dict[str, object]]]:
    grouped: dict[str, list[dict[str, object]]] = defaultdict(list)
    for match in matches:
        grouped[match["excel"]["_province"]].append(match)
    selected = {}
    for province, _ in PROVINCES:
        districts: dict[str, deque[dict[str, object]]] = defaultdict(deque)
        ordered = sorted(
            grouped[province],
            key=lambda x: (x["excel"]["_district"], x["excel"]["_level"], x["excel"]["학교명"], x["site"]["slug"]),
        )
        for match in ordered:
            districts[match["excel"]["_district"]].append(match)
        chosen = []
        district_names = sorted(districts)
        while len(chosen) < 12 and any(districts.values()):
            for district in district_names:
                if districts[district] and len(chosen) < 12:
                    chosen.append(districts[district].popleft())
        selected[province] = sorted(
            chosen, key=lambda x: (x["excel"]["_district"], x["excel"]["_level"], x["excel"]["학교명"])
        )
    return selected


def school_section(selected: dict[str, list[dict[str, object]]]) -> str:
    chips = "".join(
        f'<a href="#school-region-{code}">{html.escape(province)}</a>'
        for province, code in PROVINCES
    )
    groups = []
    for province, code in PROVINCES:
        cards = []
        for match in selected[province]:
            excel, site = match["excel"], match["site"]
            cards.append(
                f'<a class="school-card" href="/{html.escape(str(site["slug"]), quote=True)}/">'
                f'<strong class="school-name">{html.escape(excel["학교명"])}</strong>'
                f'<span class="school-location">{html.escape(excel["_level"])} · '
                f'{html.escape(excel["_district"])}</span>'
                f'<small>학교별 과외 정보 보기</small></a>'
            )
        groups.append(
            f'<section class="region-school-navigation" id="school-region-{code}">'
            f'<h3>{html.escape(province)} 학교</h3>'
            f'<div class="school-explore-grid">{"".join(cards)}</div></section>'
        )
    return (
        '<section class="home-explore-section school-explore" id="schools" aria-labelledby="school-explore-title">'
        '<div class="home-explore-heading"><div><p class="home-explore-kicker">학교 탐색</p>'
        '<h2 id="school-explore-title">학교별 과외 찾기</h2></div>'
        '<p>지역과 학교명을 기준으로 학교별 학습 정보를 살펴보세요.</p></div>'
        f'<nav class="region-school-links" aria-label="시도별 학교 탐색">{chips}</nav>'
        f'{"".join(groups)}</section>'
    )


def write_csv(path: Path, rows: list[dict[str, object]], fields: list[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    target = choose_target()
    metadata = json.loads(META.read_text(encoding="utf-8"))
    excel_rows = load_excel()
    matches, unmatched = match_site(excel_rows, metadata)
    selected = select_home(matches)
    if any(not values for values in selected.values()):
        raise RuntimeError({province: len(values) for province, values in selected.items()})

    source_home = (SOURCE / "index.html").read_text(encoding="utf-8")
    subject_start, subject_end, subject_old = balanced_section(
        source_home, '<section class="home-explore-section subject-explore"'
    )
    without_subject = source_home[:subject_start] + source_home[subject_end:]
    school_start, school_end, school_old = balanced_section(
        without_subject, '<section class="home-explore-section school-explore"'
    )
    output = without_subject[:school_start] + school_section(selected) + without_subject[school_end:]
    menu_patterns = [
        ('header_math', r'<a href="/#math">수학과외</a>'),
        ('header_english', r'<a href="/#english">영어과외</a>'),
    ]
    removed_menu = 0
    for _, pattern in menu_patterns:
        output, count = re.subn(pattern, "", output)
        removed_menu += count
    if removed_menu != 4:
        raise RuntimeError(f"expected four header/footer subject links, removed={removed_menu}")

    shutil.copytree(SOURCE, target)
    temporary = target / "index.html.tmp"
    temporary.write_text(output, encoding="utf-8", newline="")
    os.replace(temporary, target / "index.html")

    match_rows = []
    for match in matches:
        row, site = match["excel"], match["site"]
        match_rows.append({
            "excel_row": row["_row"], "school_name": row["학교명"], "sido": row["_province"],
            "sigungu": row["_district"], "school_level": row["_level"], "excel_address": row["주소"],
            "school_code": row.get("학교코드 (KEDI)", ""), "site_school_name": site.get("school_name", ""),
            "site_slug": site["slug"], "site_title": site["title"], "match_type": match["match_type"],
        })
    write_csv(MATCH_CSV, match_rows, list(match_rows[0]))
    write_csv(
        UNMATCHED_CSV, unmatched,
        ["site_school_name", "site_province", "site_sigungu", "site_address", "site_slug", "candidate_count", "reason"],
    )
    selected_slugs = {str(x["site"]["slug"]) for values in selected.values() for x in values}
    card_rows = []
    for match in matches:
        row, site = match["excel"], match["site"]
        card_rows.append({
            "sido": row["_province"], "sigungu": row["_district"], "school_name": row["학교명"],
            "school_level": row["_level"], "excel_address": row["주소"], "site_page_title": site["title"],
            "site_slug": site["slug"], "href": f'/{site["slug"]}/', "http_status": "",
            "selected_for_home": int(str(site["slug"]) in selected_slugs),
        })
    write_csv(CARD_CSV, card_rows, [
        "sido", "sigungu", "school_name", "school_level", "excel_address", "site_page_title",
        "site_slug", "href", "http_status", "selected_for_home",
    ])
    removals = [
        {"section_type": "subject_explore", "old_heading": "과목별 과외 찾기", "old_link_count": subject_old.count("<a "), "removed": 1, "reason": "홈 수학·영어 전용 탐색 제거"},
        {"section_type": "math_explore", "old_heading": "수학과외 찾기", "old_link_count": subject_old.count('data-subject="math"'), "removed": 1, "reason": "홈 수학 전용 탐색 제거"},
        {"section_type": "english_explore", "old_heading": "영어과외 찾기", "old_link_count": subject_old.count('data-subject="english"'), "removed": 1, "reason": "홈 영어 전용 탐색 제거"},
        {"section_type": "header_footer_menu", "old_heading": "수학과외·영어과외", "old_link_count": removed_menu, "removed": 1, "reason": "홈 전용 수학·영어 메뉴 제거"},
        {"section_type": "school_explore", "old_heading": "학교별 과외 찾기", "old_link_count": school_old.count("<a "), "removed": 1, "reason": "Excel 기준 학교 카드로 교체"},
    ]
    write_csv(REMOVAL_CSV, removals, ["section_type", "old_heading", "old_link_count", "removed", "reason"])

    names = Counter(norm(row["학교명"]) for row in excel_rows)
    duplicate_names = {name for name, count in names.items() if count > 1}
    checkpoint = {
        "status": "complete", "source": str(SOURCE), "target": str(target), "excel": str(EXCEL),
        "sheet": SHEET, "header_row": 17, "excel_data_rows": len(excel_rows),
        "excel_unique_school_codes": len({row.get("학교코드 (KEDI)", "") for row in excel_rows if row.get("학교코드 (KEDI)")}),
        "excel_unique_school_region_name": len({(row["_province"], row["_district"], norm(row["학교명"])) for row in excel_rows}),
        "duplicate_school_name_keys": len(duplicate_names),
        "same_name_school_rows": sum(names[name] for name in duplicate_names),
        "site_unique_schools": sum(x.get("node_type") == "school_general" for x in metadata),
        "matches": len(matches),
        "exact_matches": sum(x["match_type"] in ("exact_name_region_address", "normalized_name_region_level") for x in matches),
        "corrected_matches": sum(x["match_type"] == "address_corrected" for x in matches),
        "unmatched_site_schools": len(unmatched), "selected_count": len(selected_slugs),
        "selected_by_province": {province: len(values) for province, values in selected.items()},
        "subject_section_removed": 1, "math_section_removed": 1, "english_section_removed": 1,
        "subject_menu_links_removed": removed_menu,
    }
    CHECKPOINT.write_text(json.dumps(checkpoint, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(checkpoint, ensure_ascii=False))


if __name__ == "__main__":
    main()
